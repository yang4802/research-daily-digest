"""Append paper rows to Papers_Summary.xlsx and update 'My lack' next day.

The sheet has a header row (the row whose cells include "Title"); data rows
follow. Columns are matched by header text so the layout can move. Main-paper
rows are shaded so they stand out from reference rows.

Usage:
  # append rows from a JSON list
  python xlsx_log.py append --xlsx PATH [--sheet NAME] --json rows.json
      rows.json: [{"title","author","journal","year","summary","cons",
                   "abstract","insights","is_main":true/false}, ...]

  # fill the 'My lack' cell of a previously-logged paper (next-day update)
  python xlsx_log.py set_mylack --xlsx PATH [--sheet NAME] --match "<title substr>" --text "..."
"""
import sys, json, argparse, re
import openpyxl
from openpyxl.styles import PatternFill, Alignment

MAIN_FILL = PatternFill("solid", fgColor="FFF2CC")  # light amber for main rows
WRAP = Alignment(wrap_text=True, vertical="top")     # readable multi-line cells
WRAP_FIELDS = ("summary", "cons", "abstract", "insights")

def fmt_numbered(text):
    """Put each '1. 2. 3.' point on its own line inside the cell. Decimals like
    '13.56' or '93.5%' are untouched (no space after the dot)."""
    if not isinstance(text, str):
        return text
    return re.sub(r'\s*(\d+)\.\s+', lambda m: ('\n' if m.start() else '') + m.group(1) + '. ', text).strip()

def abstract_from_pdf(pdf, limit=1600):
    """Pull the abstract text (between 'Abstract' and 'Index Terms'/'Introduction')
    from a PDF's first two pages. Returns '' on failure."""
    try:
        from pypdf import PdfReader
        r = PdfReader(pdf)
        txt = " ".join((r.pages[i].extract_text() or "") for i in range(min(2, len(r.pages)))).replace("\n", " ")
    except Exception:
        return ""
    m = re.search(r'Abstract\s*[—–\-:]*\s*(.*?)(Index\s*Terms|I\.?\s*I?\s*NTRODUCTION|Keywords)', txt, re.I)
    s = m.group(1) if m else txt[:limit]
    return re.sub(r'\s+', ' ', s).strip()[:limit]

# header text -> our field key
HEADER_MAP = {
    "title": "title", "author": "author", "journal name": "journal",
    "journal": "journal", "year": "year", "short summary": "summary",
    "cons of the paper": "cons", "cons": "cons", "my lack": "mylack",
    "abstract": "abstract", "insights": "insights",
}

def find_layout(ws):
    """Return (header_row, {field: col_idx}, num_col_idx)."""
    for r in range(1, 15):
        cols, numcol = {}, None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = str(v).strip().lower()
            if key == "#":
                numcol = c
            elif key in HEADER_MAP:
                cols[HEADER_MAP[key]] = c
        if "title" in cols:           # this is the header row
            return r, cols, numcol
    raise SystemExit("Could not find a header row containing 'Title'.")

def first_empty_title_row(ws, header_row, title_col):
    """First row (after header) whose Title cell is empty. The '#' column is a
    drag-down formula (=IF(COUNTA(Bx)>0,ROW()-4,"")), so emptiness is judged by
    Title only — never by the '#' column."""
    r = header_row + 1
    while ws.cell(r, title_col).value not in (None, ""):
        r += 1
    return r

def cmd_append(args):
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb[args.sheet] if args.sheet else wb.active
    hr, cols, numcol = find_layout(ws)
    rows = json.load(open(args.json, encoding="utf-8"))
    r = first_empty_title_row(ws, hr, cols["title"])
    for item in rows:
        if not item.get("abstract") and item.get("abstract_pdf"):
            item["abstract"] = abstract_from_pdf(item["abstract_pdf"])
        for field in ("title", "author", "journal", "year", "summary", "cons", "abstract", "insights"):
            if field in cols and item.get(field) is not None:
                val = fmt_numbered(item[field]) if field == "summary" else item[field]
                cell = ws.cell(r, cols[field], val)
                if field in WRAP_FIELDS:
                    cell.alignment = WRAP
        if "mylack" in cols:
            ws.cell(r, cols["mylack"], item.get("mylack", ""))  # blank now; filled next day
        # '#' is a drag-down formula. Preserve it where present; add it only if missing.
        if numcol and ws.cell(r, numcol).value in (None, ""):
            ws.cell(r, numcol, f'=IF(COUNTA(B{r})>0,ROW()-4,"")')
        if item.get("is_main"):
            lo = min([numcol or 1] + list(cols.values()))
            hi = max([numcol or 1] + list(cols.values()))
            for c in range(lo, hi + 1):
                ws.cell(r, c).fill = MAIN_FILL
        print(f"  + row {r} {'[MAIN]' if item.get('is_main') else 'ref  '} {str(item.get('title',''))[:60]}")
        r += 1
    wb.save(args.xlsx)
    print(f"Saved {args.xlsx}")

def cmd_set_mylack(args):
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb[args.sheet] if args.sheet else wb.active
    hr, cols, numcol = find_layout(ws)
    if "mylack" not in cols:
        raise SystemExit("No 'My lack' column found.")
    tcol = cols["title"]
    hits = 0
    for r in range(hr + 1, ws.max_row + 1):
        title = ws.cell(r, tcol).value
        if title and args.match.lower() in str(title).lower():
            ws.cell(r, cols["mylack"], args.text); hits += 1
            print(f"  set My lack @row {r}: {str(title)[:50]}")
    if not hits:
        print(f"No row matched '{args.match}'.")
    wb.save(args.xlsx)

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    a = sub.add_parser("append"); a.add_argument("--xlsx", required=True); a.add_argument("--sheet", default=None); a.add_argument("--json", required=True)
    m = sub.add_parser("set_mylack"); m.add_argument("--xlsx", required=True); m.add_argument("--sheet", default=None); m.add_argument("--match", required=True); m.add_argument("--text", required=True)
    args = ap.parse_args()
    {"append": cmd_append, "set_mylack": cmd_set_mylack}[args.cmd](args)

if __name__ == "__main__":
    main()
